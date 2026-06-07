#!/usr/bin/env python3
# coding: utf-8
"""
MAPEAR VALORES REAIS — lê com data_only=True
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

TEMPLATE_PATH = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO IV Planilha Orçamentária.xlsx"

print("=" * 130)
print("📊 MAPA DE ITENS — VALORES REAIS")
print("=" * 130)

# Ler com data_only=True para ver valores calculados
try:
    wb_values = load_workbook(TEMPLATE_PATH, data_only=True)
    ws_val = wb_values.active

    # Ler com data_only=False para ver fórmulas (referência)
    wb_formulas = load_workbook(TEMPLATE_PATH, data_only=False)
    ws_form = wb_formulas.active

    print(f"\nPlanilha: {ws_val.title}")
    print(f"Cabeçalho esperado: Linha 20")
    print(f"Primeiros itens: Linha 21+\n")

    # Mostrar cabeçalho
    print("📋 CABEÇALHO (Linha 20):")
    for col_idx in range(1, 14):
        cell = ws_val.cell(row=20, column=col_idx)
        col_letter = chr(64 + col_idx)  # A=65...
        print(f"   {col_letter}: {str(cell.value)[:45] if cell.value else '---'}")

    # Encontrar itens com dados (não-vazios)
    print("\n📝 ITENS COM DADOS:")
    items_found = []

    for row_idx in range(21, 100):  # Primeiros 80 itens
        item_num = ws_val.cell(row=row_idx, column=1).value
        etapa = ws_val.cell(row=row_idx, column=2).value
        descricao = ws_val.cell(row=row_idx, column=3).value

        # Item válido se tem número (1-50) e etapa ou descrição
        if isinstance(item_num, (int, float)) and item_num >= 1 and item_num <= 50:
            if etapa or descricao:
                # Coletar todos os valores de interesse
                justif_proponent = ws_val.cell(row=row_idx, column=4).value
                justif_terceiro = ws_val.cell(row=row_idx, column=5).value
                unidade = ws_val.cell(row=row_idx, column=8).value
                quantidade = ws_val.cell(row=row_idx, column=9).value
                especificacao = ws_val.cell(row=row_idx, column=10).value
                valor_unitario = ws_val.cell(row=row_idx, column=11).value
                valor_total = ws_val.cell(row=row_idx, column=12).value
                provimento = ws_val.cell(row=row_idx, column=13).value

                items_found.append({
                    "num": int(item_num),
                    "linha": row_idx,
                    "etapa": etapa,
                    "descricao": descricao,
                    "unidade": unidade,
                    "quantidade": quantidade,
                    "valor_unitario": valor_unitario,
                    "valor_total": valor_total,
                    "provimento": provimento
                })

    # Mostrar items encontrados
    for item in items_found[:40]:
        print(f"\n   🔹 ITEM {item['num']} (Linha {item['linha']})")
        print(f"      Etapa: {item['etapa']}")
        print(f"      Descrição: {str(item['descricao'])[:60]}")
        if item['unidade']:
            print(f"      Unidade: {item['unidade']} | Qtd: {item['quantidade']}")
        if item['valor_unitario'] is not None:
            print(f"      Valor Unit: {item['valor_unitario']:>12} | Total: {item['valor_total']:>12}")

    # Criar tabela de alterações
    print("\n" + "=" * 130)
    print("🎯 MAPA DE ALTERAÇÕES")
    print("=" * 130)

    alteracoes = {
        2: {"campo": "valor_total", "novo_valor": 8000, "descricao": "Direção Criação"},
        4: {"campo": "desc", "novo_valor": "Direção de Arte - Conceituação Artística", "descricao": "Sinografia → Dir. Arte"},
        4: {"campo": "valor_total", "novo_valor": 4000},
        7: {"campo": "valor_total", "novo_valor": 6000, "descricao": "Figurinos"},
        10: {"campo": "valor_total", "novo_valor": 8000, "descricao": "Dir. Montagem"},
        11: {"campo": "quantidade", "novo_valor": 2, "descricao": "Assistência (1mês→2mês)"},
        11: {"campo": "valor_unitario", "novo_valor": 2000},
        11: {"campo": "valor_total", "novo_valor": 4000},
        14: {"campo": "desc", "novo_valor": "Confecção de Cenografia + Adereços", "descricao": "Cenografia"},
        14: {"campo": "valor_total", "novo_valor": 12000},
        19: {"campo": "valor_unitario", "novo_valor": 600, "descricao": "Cachê apresentação"},
        19: {"campo": "valor_total", "novo_valor": 3600},
    }

    print("\nPronto para aplicar alterações:")
    for item in items_found:
        if item['num'] in [2, 4, 7, 10, 11, 14, 19]:
            print(f"\n   ✏️  Item {item['num']} (Linha {item['linha']})")
            print(f"      Atual: {str(item['descricao'])[:50]} | Total: {item['valor_total']}")

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
