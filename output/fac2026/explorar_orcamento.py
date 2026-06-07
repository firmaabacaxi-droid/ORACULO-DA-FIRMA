#!/usr/bin/env python3
# coding: utf-8
"""
EXPLORAÇÃO DETALHADA — encontrar estrutura de itens
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

TEMPLATE_PATH = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO IV Planilha Orçamentária.xlsx"

print("=" * 100)
print("🔎 EXPLORAÇÃO DETALHADA")
print("=" * 100)

try:
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb.active

    print(f"\nPlanilha: {ws.title}")
    print(f"Dimensões: {ws.dimensions}")
    print(f"Máx coluna: {ws.max_column}, Máx linha: {ws.max_row}")

    # Procurar por padrões: "Item 1", "1.", números na primeira coluna, etc
    print("\n📊 PROCURANDO PADRÕES DE ITENS...")

    print("\n➤ LINHAS 1-50:")
    for row_idx in range(1, 51):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value

        # Mostrar se tem conteúdo interessante
        if any([col_a, col_b, col_c]):
            a_str = str(col_a)[:20] if col_a else "---"
            b_str = str(col_b)[:20] if col_b else "---"
            c_str = str(col_c)[:30] if col_c else "---"
            print(f"  [{row_idx:3d}] A:{a_str:<20} | B:{b_str:<20} | C:{c_str:<30}")

    # Procurar números iniciais nas colunas A ou B
    print("\n➤ PROCURANDO NÚMEROS (possíveis números de item):")
    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        if isinstance(col_a, (int, float)) and 1 <= col_a <= 50:
            col_c = ws.cell(row=row_idx, column=3).value
            col_l = ws.cell(row=row_idx, column=12).value
            print(f"  [Linha {row_idx:3d}] Item {col_a:2.0f}: {str(col_c)[:40] if col_c else '---':<40} | Total: {col_l}")

    # Procurar por palavras-chave
    print("\n➤ PROCURANDO PALAVRAS-CHAVE:")
    keywords = ["direção", "figurino", "cenografia", "videoprojeção", "iluminação", "total", "r$"]
    for row_idx in range(1, ws.max_row + 1):
        row_str = " ".join(str(ws.cell(row=row_idx, column=col).value or "") for col in range(1, 13)).lower()
        if any(kw in row_str for kw in keywords):
            if row_idx <= 100 or "total" in row_str:
                content = " | ".join(str(ws.cell(row=row_idx, column=col).value or "")[:15] for col in range(1, 8))
                print(f"  [Linha {row_idx:3d}] {content}")

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
