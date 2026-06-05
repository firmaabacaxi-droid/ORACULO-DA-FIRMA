"""
Preenche o Anexo IX (Cronograma de Execução) do FAC 2026
seguindo a estrutura do template ODS oficial.

Fonte dos dados: ANEXO-IX-Cronograma-Todas-as-Historias-do-Mundo.xlsx
Template: ANEXO IX - CRONOGRAMA DE EXEUÇÃO.ods

Normalização de meses:
  XLSX "Mês 2" → Projeto Mês 1
  XLSX "Mês 3" → Projeto Mês 2
  ... (subtrai 1 do número do mês)

Saída: ANEXO-IX-Cronograma-Todas-as-Historias-do-Mundo-TEMPLATE.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter

# ── DADOS DO CRONOGRAMA ──────────────────────────────────────────────────────
# (atividade, descrição, local, início_mês, início_sem, término_mês, término_sem)

PRE_PRODUCAO = [
    ("Concepção e criação dramatúrgica",
     "Pesquisa de referências, escrita e estruturação do roteiro",
     "Brasília - DF", 1, 1, 2, 4),
    ("Direção e pesquisa",
     "Desenvolvimento conceitual, pesquisa artística e direção do processo criativo",
     "Brasília - DF", 1, 1, 2, 4),
    ("Preparação corporal e vocal",
     "Treinamento de técnicas circenses e preparação vocal do intérprete",
     "Brasília - DF", 1, 1, 2, 4),
    ("Criação de conteúdo videoprojeção",
     "Desenvolvimento de vídeos e mapeamento de projeção nos tecidos",
     "Brasília - DF", 1, 1, 2, 4),
    ("Projeto cenografia e iluminação",
     "Concepção e detalhamento técnico da cenografia e design de luz",
     "Brasília - DF", 1, 1, 2, 4),
    ("Trilha sonora e desenho de som",
     "Composição musical e criação de paisagem sonora original",
     "Brasília - DF", 1, 1, 2, 4),
    ("Criação de figurinos",
     "Pesquisa, projeto e confecção de figurinos do espetáculo",
     "Brasília - DF", 1, 1, 2, 4),
    ("Operador de cena (ensaios)",
     "Operação técnica nas sessões de ensaio e montagem",
     "Brasília - DF", 2, 1, 3, 4),
    ("Aluguel de sala de ensaio",
     "Locação de espaço para ensaios técnicos e artísticos",
     "Brasília - DF", 2, 1, 3, 4),
]

PRODUCAO = [
    ("Ensaios e montagem (atuação)",
     "Ensaios técnicos e artísticos, integração de todos os elementos cênicos",
     "Brasília - DF", 2, 1, 3, 4),
    ("Assistência de direção",
     "Apoio à direção artística nos ensaios e montagem final",
     "Brasília - DF", 2, 1, 3, 4),
    ("Confecção cenografia, figurinos, adereços",
     "Construção e finalização de todos os elementos cênicos",
     "Brasília - DF", 2, 1, 3, 4),
    ("Locação de equipamentos",
     "Aluguel de rigging, projetores, trilhos e equipamentos circenses",
     "Brasília - DF", 3, 3, 3, 4),
    ("Montagem e testes técnicos",
     "Montagem em teatro e testes integrados de luz, som e projeção mapeada",
     "Brasília - DF", 3, 3, 3, 4),
    ("Apresentação 1 — Estreia",
     "Estreia do espetáculo para público geral (semana 1)",
     "Brasília - DF", 3, 1, 3, 1),
    ("Apresentações 2 e 3",
     "Segunda e terceira sessões do espetáculo (semana 2)",
     "Brasília - DF", 3, 2, 3, 2),
    ("Apresentações 4 e 5",
     "Quarta e quinta sessões do espetáculo (semana 3)",
     "Brasília - DF", 3, 3, 3, 3),
    ("Apresentação 6 + Online",
     "Última sessão presencial e transmissão ao vivo online acessível (semana 4)",
     "Brasília - DF", 3, 4, 3, 4),
    ("Audiodescrição e acessibilidade",
     "Sessão acessível com audiodescrição, Libras e recursos de inclusão",
     "Brasília - DF", 3, 1, 3, 4),
]

POS_PRODUCAO = [
    ("Registro audiovisual e fotografia",
     "Captação profissional de imagens e vídeos do espetáculo em cena",
     "Brasília - DF", 3, 1, 4, 4),
    ("Edição e finalização audiovisual",
     "Montagem, correção de cor, mixagem e entrega do material audiovisual",
     "Brasília - DF", 4, 1, 5, 4),
    ("Divulgação (redes, imprensa, assessoria)",
     "Campanha em redes sociais, assessoria de imprensa e divulgação na mídia",
     "Brasília - DF", 4, 1, 6, 4),
    ("Gestão e fechamento do projeto",
     "Administração, prestação de contas e elaboração do relatório final",
     "Brasília - DF", 5, 1, 6, 4),
]

# ── ESTILOS ──────────────────────────────────────────────────────────────────

CINZA_ESCURO  = "2F4F4F"   # cabeçalho principal
CINZA_MEDIO   = "4A7C7C"   # seções
CINZA_CLARO   = "D9E5E5"   # linhas alternadas
VERDE_FAC     = "006666"   # destaques
BRANCO        = "FFFFFF"

def bold_white_center(sz=10):
    return Font(bold=True, color=BRANCO, size=sz)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def borda():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── CRIAÇÃO DA PLANILHA ──────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cronograma"

# Larguras de coluna
col_widths = {
    "A": 5,    # Nº
    "B": 38,   # Atividade Geral
    "C": 42,   # Descrição
    "D": 16,   # Local (RA/UF/Espaço)
    "E": 10,   # Início Mês
    "F": 12,   # Início Semana
    "G": 10,   # Término Mês
    "H": 12,   # Término Semana
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── LINHA 1: CABEÇALHO INSTITUCIONAL ────────────────────────────────────────
ws.merge_cells("A1:H1")
ws["A1"] = "GOVERNO DO DISTRITO FEDERAL — SECRETARIA DE ESTADO DE CULTURA E ECONOMIA CRIATIVA"
ws["A1"].font = bold_white_center(11)
ws["A1"].fill = fill(CINZA_ESCURO)
ws["A1"].alignment = align("center")
ws.row_dimensions[1].height = 22

# ── LINHA 2: TÍTULO DO ANEXO ─────────────────────────────────────────────────
ws.merge_cells("A2:H2")
ws["A2"] = "ANEXO IX — CRONOGRAMA DE EXECUÇÃO"
ws["A2"].font = bold_white_center(12)
ws["A2"].fill = fill(VERDE_FAC)
ws["A2"].alignment = align("center")
ws.row_dimensions[2].height = 24

# ── LINHAS 3-5: DADOS DO PROJETO ─────────────────────────────────────────────
info_font = Font(bold=True, size=10)
val_font  = Font(size=10)

def info_row(ws, row, label, value):
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = info_font
    ws[f"A{row}"].alignment = align()
    ws.merge_cells(f"C{row}:H{row}")
    ws[f"C{row}"] = value
    ws[f"C{row}"].font = val_font
    ws[f"C{row}"].alignment = align()
    ws.row_dimensions[row].height = 18
    for col in ["A", "C"]:
        ws[f"{col}{row}"].fill = fill("F5F5F5")

info_row(ws, 3, "PROJETO", "Todas as Histórias do Mundo")
info_row(ws, 4, "PROPONENTE", "Filipe Duque")
info_row(ws, 5, "PERÍODO DE EXECUÇÃO", "6 meses (data de início a confirmar após aprovação)")

# ── LINHA 6: CABEÇALHOS DAS COLUNAS ──────────────────────────────────────────
headers = [
    ("A6", "Nº"),
    ("B6", "Atividade Geral\n(realização, investigação, gravação, etc.)"),
    ("C6", "Descrição\n(contratação, ensaio, oficina, impressão, etc.)"),
    ("D6", "Local\n(RA, UF, espaço físico)"),
    ("E6", "Início\nMês"),
    ("F6", "Início\nSemana"),
    ("G6", "Término\nMês"),
    ("H6", "Término\nSemana"),
]
for cell_ref, text in headers:
    cell = ws[cell_ref]
    cell.value = text
    cell.font = bold_white_center(9)
    cell.fill = fill(CINZA_ESCURO)
    cell.alignment = align("center")
    cell.border = borda()
ws.row_dimensions[6].height = 30

# ── FUNÇÃO: SEÇÃO ─────────────────────────────────────────────────────────────
def write_section(ws, start_row, titulo, atividades, counter_start):
    # Linha de seção
    ws.merge_cells(f"A{start_row}:H{start_row}")
    ws[f"A{start_row}"] = titulo
    ws[f"A{start_row}"].font = bold_white_center(10)
    ws[f"A{start_row}"].fill = fill(CINZA_MEDIO)
    ws[f"A{start_row}"].alignment = align("left")
    ws.row_dimensions[start_row].height = 20

    row = start_row + 1
    counter = counter_start

    for idx, (ativ, desc, local, ini_mes, ini_sem, term_mes, term_sem) in enumerate(atividades):
        bg = CINZA_CLARO if idx % 2 == 0 else BRANCO
        data = [counter, ativ, desc, local, ini_mes, ini_sem, term_mes, term_sem]
        cols  = ["A", "B", "C", "D", "E", "F", "G", "H"]
        aligns = ["center", "left", "left", "left", "center", "center", "center", "center"]

        for col, val, ali in zip(cols, data, aligns):
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.fill = fill(bg)
            cell.border = borda()
            cell.alignment = align(ali)
            if col == "A":
                cell.font = Font(bold=True, size=9)
            elif col in ("E", "F", "G", "H"):
                cell.font = Font(size=9, bold=True, color=VERDE_FAC)
            else:
                cell.font = Font(size=9)
        ws.row_dimensions[row].height = 28
        row += 1
        counter += 1

    return row, counter  # próxima linha disponível, próximo Nº


# ── ESCREVER AS 3 SEÇÕES ──────────────────────────────────────────────────────
next_row, next_num = write_section(
    ws, 7, "Pré-Produção / Preparação", PRE_PRODUCAO, 1
)
next_row, next_num = write_section(
    ws, next_row, "Produção / Realização", PRODUCAO, next_num
)
next_row, next_num = write_section(
    ws, next_row, "Pós-Produção / Finalização", POS_PRODUCAO, next_num
)

# ── LEGENDA ───────────────────────────────────────────────────────────────────
next_row += 1
ws.merge_cells(f"A{next_row}:H{next_row}")
leg = ws[f"A{next_row}"]
leg.value = (
    "Nota: semanas numeradas de 1 a 4 dentro de cada mês. "
    "Meses numerados sequencialmente a partir do início do projeto (Mês 1 = primeiro mês de execução)."
)
leg.font = Font(italic=True, size=8, color="555555")
leg.alignment = align("left", wrap=True)
ws.row_dimensions[next_row].height = 28

# ── SALVAR ────────────────────────────────────────────────────────────────────
OUTPUT = (
    r"PROJETOS - EDITAIS\EDITAL FAC 2026"
    r"\ANEXO-IX-Cronograma-Todas-as-Historias-do-Mundo-TEMPLATE.xlsx"
)
wb.save(OUTPUT)
print(f"Arquivo salvo: {OUTPUT}")
print(f"Total de atividades: {next_num - 1}")
