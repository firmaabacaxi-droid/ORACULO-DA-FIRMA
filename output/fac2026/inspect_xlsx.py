import openpyxl

wb = openpyxl.load_workbook(r'PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO-IX-Cronograma-Todas-as-Historias-do-Mundo.xlsx')
print(f'Sheets: {wb.sheetnames}')

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== SHEET: {name} (max_row={ws.max_row}, max_col={ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        # Remover linhas totalmente vazias
        if any(c is not None and str(c).strip() != '' for c in row):
            print(f'  {list(row)}')
