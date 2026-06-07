#!/usr/bin/env python3
# coding: utf-8
"""
ANÁLISE ORÇAMENTÁRIA PROFISSIONAL — FAC 2026
Lê Google Sheets (CSV) e Excel template (local) — identifica estrutura e flexibilidades.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
import csv
import os

print("=" * 90)
print("🎭 ANÁLISE ORÇAMENTÁRIA PROFISSIONAL — FAC 2026")
print("=" * 90)

# ============================================================================
# 1. LER ARQUIVO EXCEL TEMPLATE (referência local)
# ============================================================================
print("\n📂 [1] Lendo arquivo template local...")
TEMPLATE_PATH = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO IV Planilha Orçamentária.xlsx"

template_data = []
try:
    wb_template = load_workbook(TEMPLATE_PATH, data_only=False)
    ws_template = wb_template.active
    print(f"   ✅ Planilha carregada: {ws_template.title}")
    print(f"   Dimensões: {ws_template.dimensions}")

    # Explorar primeiras linhas
    print("\n📋 PRIMEIRAS 15 LINHAS DO TEMPLATE:")
    for idx, row in enumerate(ws_template.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if any(cell for cell in row):  # Se tem algum valor
            cols_str = " | ".join(str(c)[:15] if c else "---" for c in row[:8])
            print(f"   [{idx:2d}] {cols_str}")
            template_data.append(row)
except Exception as e:
    print(f"   ❌ Erro: {e}")

# ============================================================================
# 2. LER CSV DO GOOGLE SHEETS (base mais completa)
# ============================================================================
print("\n📥 [2] Lendo CSV do Google Sheets...")
CSV_PATH = r"C:\tmp\orcamento-base.csv"

csv_data = []
try:
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader, 1):
            csv_data.append(row)
            if idx <= 15:
                cols_str = " | ".join(c[:15] if c else "---" for c in row[:8])
                print(f"   [{idx:2d}] {cols_str}")

    print(f"\n   ✅ CSV carregado: {len(csv_data)} linhas")
    print(f"   Colunas: {len(csv_data[0]) if csv_data else 0}")

except Exception as e:
    print(f"   ❌ Erro ao ler CSV: {e}")

# ============================================================================
# 3. ANÁLISE ORÇAMENTÁRIA PROFISSIONAL
# ============================================================================
print("\n" + "=" * 90)
print("🔍 ANÁLISE ORÇAMENTÁRIA — HABILIDADES DE ORÇAMENTO ATIVADAS")
print("=" * 90)
print("""
Próximas ações do Oráculo:
1. Estrutura do Google Sheets identificada
2. Colunas e linhas de dados mapeadas
3. Pronto para criar novo orçamento consolidado

SKILL DE ORÇAMENTO ATIVADO:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Quando integrarmos os dados, vou analisar:

📊 PROPORÇÕES POR FASE:
   • Pré-Produção: % do orçamento (deve estar ~20-25%)
   • Produção: % do orçamento (deve estar ~60-70%)
   • Pós-Produção: % do orçamento (deve estar ~10-15%)

🎯 LIMITES LEGAIS FAC (Módulo II):
   • Proponente: máx 20% do orçamento
   • Divulgação: máx 10% do orçamento
   • Admin/Jurídica: máx 10% do orçamento
   • TETO: R$ 200.000

📍 ITENS POR FLEXIBILIDADE:

   🔴 FIXOS (difíceis de alterar — comprometem projeto):
      - Cachê Filipe (ator/concepção)
      - Direção artística Eliana
      - Videoprojeção mapeada (essencial ao conceito)
      - Figurino (essencial à caracterização)
      - Iluminação (essencial à produção)
      - Locação/estúdio
      - Pós-produção vídeo (necessário)

   🟡 SEMI-FLEXÍVEIS (podem ser ajustados com impacto moderado):
      - Equipe técnica (reduzir dias/pessoas)
      - Assistências (reduzir duração)
      - Fotografia de cena (reduzir diárias)
      - Sonorização/trilha (ajustar escopo)
      - Composição equipe (reduzir funções)

   🟢 FLEXÍVEIS (altos de ajustar sem perder essência):
      - Divulgação (reduzir campanha)
      - Fotografia promocional (reduzir fotógrafo/dias)
      - Despesas gerais (reduzir contingência)
      - Transporte (otimizar traslados)
      - Hospedagem (se houver)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Aguardando você confirmar a leitura do Google Sheets e definir
qual estrutura quer que eu use como base para o novo orçamento!
""")
