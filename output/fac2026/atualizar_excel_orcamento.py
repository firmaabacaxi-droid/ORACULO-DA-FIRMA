#!/usr/bin/env python3
# coding: utf-8
import sys
import os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from datetime import datetime

# Caminhos
ARQUIVO_ORIGINAL = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO-IV-Planilha-Orcamentaria-Todas-as-Historias-do-Mundo.xlsx"
ARQUIVO_NOVO = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\output\fac2026\ANEXO-IV-REVISADO.xlsx"
PASTA_DRIVE = "1siKUTBMJxXJywPXCAsvZA5G5L8xcoirM"

print("[INICIANDO] Atualizacao do orcamento FAC 2026")
print("="*80)

try:
    print("\n[PROCESSANDO] Carregando planilha original...")
    wb = load_workbook(ARQUIVO_ORIGINAL)
    ws = wb.active
    print(f"[OK] Planilha carregada: {ws.title}")

    print("\n[PROCESSANDO] Aplicando alteracoes...")

    # Mapeamento de alteracoes (linha, coluna, novo_valor)
    # Estrutura da planilha: A=1, B=2, C=3, ..., J=10, K=11, etc
    alteracoes = [
        # Item 2 (linha ~23): Direcao criacao R$ 10.000 -> R$ 8.000
        (23, 10, "8000"),    # K23 - Valor unitario
        (23, 11, "8000"),    # L23 - Valor total

        # Item 4 (linha ~25): Projeto cenografia -> Direcao de Arte, R$ 3.000 -> R$ 4.000
        (25, 3, "Direcao de Arte - Concepcao Artistica"),  # C25
        (25, 10, "4000"),    # K25
        (25, 11, "4000"),    # L25

        # Item 7 (linha ~28): Criacao figurinos R$ 4.000 -> R$ 6.000
        (28, 10, "6000"),    # K28
        (28, 11, "6000"),    # L28

        # Item 10 (linha ~31): Direcao montagem R$ 10.000 -> R$ 8.000
        (31, 10, "8000"),    # K31
        (31, 11, "8000"),    # L31

        # Item 11 (linha ~32): Assistencia direcao 1 mes -> 2 meses, R$ 2.000 -> R$ 4.000
        (32, 8, 2),          # I32 - Quantidade
        (32, 10, "2000"),    # K32 - Valor unitario
        (32, 11, "4000"),    # L32 - Valor total

        # Item 14 (linha ~35): Aderecos -> Confeccao cenografia + aderecos, R$ 2.000 -> R$ 12.000
        (35, 3, "Confeccao Cenografia + Aderecos e Objetos"),  # C35
        (35, 10, "12000"),   # K35
        (35, 11, "12000"),   # L35

        # Item 19 (linha ~40): Cache apresentacao R$ 2.000 -> R$ 600, total R$ 12.000 -> R$ 3.600
        (40, 10, "600"),     # K40
        (40, 11, "3600"),    # L40
    ]

    for linha, col, valor in alteracoes:
        try:
            celula = ws.cell(row=linha, column=col)
            celula.value = valor
            print(f"  [OK] Linha {linha}, Col {col}: {valor}")
        except Exception as e:
            print(f"  [ERRO] Linha {linha}: {e}")

    print("\n[PROCESSANDO] Removendo itens consolidados...")
    # Remover linhas de itens consolidados (em ordem reversa)
    try:
        ws.delete_rows(34, 1)  # Item 12
        print("  [OK] Item 12 removido")
        ws.delete_rows(29, 1)  # Item 8
        print("  [OK] Item 8 removido")
    except Exception as e:
        print(f"  [AVISO] Nao foi possivel remover linhas: {e}")

    print("\n[PROCESSANDO] Salvando nova versao...")
    wb.save(ARQUIVO_NOVO)
    print(f"[OK] Arquivo salvo: {ARQUIVO_NOVO}")

    print("\n[PROCESSANDO] Fazendo upload para Google Drive...")
    import subprocess
    resultado = subprocess.run(
        ["rclone", "copy", ARQUIVO_NOVO, f"gdrive:{PASTA_DRIVE}/"],
        capture_output=True,
        text=True
    )
    if resultado.returncode == 0:
        print("[OK] Upload concluido no Drive")
    else:
        print(f"[AVISO] Upload: {resultado.stderr}")

    print("\n" + "="*80)
    print("[SUCESSO] Orcamento atualizado e enviado!")
    print(f"\nArquivo local: {ARQUIVO_NOVO}")
    print(f"Drive: https://drive.google.com/drive/folders/{PASTA_DRIVE}")
    print("\nProxima etapa: Abra o arquivo ANEXO-IV-REVISADO.xlsx no Drive")
    print("e verifique o TOTAL no final da planilha")

except Exception as e:
    print(f"\n[ERRO FATAL] {e}")
    import traceback
    traceback.print_exc()
