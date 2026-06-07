#!/usr/bin/env python3
# coding: utf-8
"""
MAPEAR ESTRUTURA DO ORÇAMENTO
Identifica exatamente linhas, colunas e valores de cada item.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = r"C:\Users\User\Documents\ORACULO - FIRMA ABACAXI\PROJETOS - EDITAIS\EDITAL FAC 2026\ANEXO IV Planilha Orçamentária.xlsx"

print("=" * 110)
print("🗺️ MAPA ESTRUTURAL — ORÇAMENTO FAC 2026")
print("=" * 110)

try:
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb.active

    # Procurar a linha que começa com "ITEM"
    item_header_row = None
    for row_idx in range(1, 100):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a and "ITEM" in str(cell_a).upper():
            item_header_row = row_idx
            break

    print(f"\n📍 Linha de cabeçalho: {item_header_row}")

    if item_header_row:
        # Mostrar cabeçalhos
        print("\n📊 COLUNAS:")
        cols_map = {}
        for col_idx in range(1, 14):
            cell = ws.cell(row=item_header_row, column=col_idx)
            col_letter = get_column_letter(col_idx)
            value = cell.value
            cols_map[col_idx] = value
            if value:
                print(f"   {col_letter:3} ({col_idx:2}) = {str(value)[:60]}")

        # Encontrar itens
        print(f"\n📝 ITENS ENCONTRADOS (começando linha {item_header_row + 1}):")

        items = {}
        for row_idx in range(item_header_row + 1, item_header_row + 50):
            item_num = ws.cell(row=row_idx, column=1).value
            item_desc = ws.cell(row=row_idx, column=3).value
            item_un = ws.cell(row=row_idx, column=8).value
            item_qtd = ws.cell(row=row_idx, column=9).value
            item_val_unit = ws.cell(row=row_idx, column=11).value
            item_val_total = ws.cell(row=row_idx, column=12).value

            # Se tem número de item, é um item válido
            if item_num and isinstance(item_num, (int, float)):
                items[int(item_num)] = {
                    "linha": row_idx,
                    "descricao": item_desc,
                    "unidade": item_un,
                    "quantidade": item_qtd,
                    "valor_unitario": item_val_unit,
                    "valor_total": item_val_total
                }

                print(f"\n   🔹 ITEM {int(item_num)} — Linha {row_idx}")
                print(f"      Descrição: {item_desc}")
                print(f"      Unidade: {item_un} | Qtd: {item_qtd}")
                print(f"      Valor unitário: {item_val_unit}")
                print(f"      Valor total: {item_val_total}")

        # Mostrar mapa resumido
        print("\n" + "=" * 110)
        print("📋 MAPA RESUMIDO — LOCALIZAÇÃO DOS ITENS")
        print("=" * 110)
        print(f"{'Item':<6} {'Linha':<8} {'Descrição':<50} {'Colunas':<30}")
        print("-" * 110)

        for num in sorted(items.keys()):
            info = items[num]
            desc = str(info['descricao'])[:48] if info['descricao'] else "---"
            cols_info = f"Desc:C | Un:H | Val.Unit:K | Val.Total:L"
            print(f"{num:<6} {info['linha']:<8} {desc:<50} {cols_info:<30}")

        # Mapa de alterações por item
        print("\n" + "=" * 110)
        print("🎯 MAPA DE ALTERAÇÕES")
        print("=" * 110)

        alteracoes_map = {
            2: {"desc_novo": "Direção e Conceituação", "val_novo": 8000},
            4: {"desc_novo": "Direção de Arte - Conceituação Artística", "val_novo": 4000},
            7: {"desc_novo": "Criação de Figurinos", "val_novo": 6000},
            10: {"desc_novo": "Direção de Montagem", "val_novo": 8000},
            11: {"qtd_novo": 2, "val_unit_novo": 2000, "val_total_novo": 4000},
            14: {"desc_novo": "Confecção de Cenografia + Adereços", "val_novo": 12000},
            19: {"val_unit_novo": 600, "val_total_novo": 3600}
        }

        for item_num, changes in alteracoes_map.items():
            if item_num in items:
                info = items[item_num]
                print(f"\n✏️  Item {item_num} (Linha {info['linha']})")
                for key, val in changes.items():
                    if key == "desc_novo":
                        col_letra = "C"
                        print(f"   {col_letra} = '{val}'")
                    elif key == "val_novo":
                        col_letra = "L"
                        print(f"   {col_letra} = {val}")
                    elif key == "val_unit_novo":
                        col_letra = "K"
                        print(f"   {col_letra} = {val}")
                    elif key == "val_total_novo":
                        col_letra = "L"
                        print(f"   {col_letra} = {val}")
                    elif key == "qtd_novo":
                        col_letra = "I"
                        print(f"   {col_letra} = {val}")

except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()
